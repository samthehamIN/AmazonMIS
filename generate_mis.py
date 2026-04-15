#!/usr/bin/env python3
"""
PG Beauty Amazon MIS Generator
Reads all weekly Excel scorecards and generates a comprehensive HTML MIS.
Run this script each time a new weekly Excel is added to the folder.
"""

import os
import re
import json
import math
from glob import glob
import openpyxl
from datetime import datetime, timedelta

# ──────────────────────────────────────────────
# WEEK DATE MAP  (week-ending Saturday dates)
# ──────────────────────────────────────────────
WEEK_DATES = {
    2:  "2026-01-10",
    9:  "2026-02-28",
    10: "2026-03-07",
    11: "2026-03-14",
    12: "2026-03-21",
    13: "2026-03-28",
    14: "2026-04-04",
    15: "2026-04-11",
}

# ──────────────────────────────────────────────
# PARSERS
# ──────────────────────────────────────────────

def safe_float(v):
    try:
        if v is None or v == '' or v == '#DIV/0!':
            return None
        return float(v)
    except (ValueError, TypeError):
        return None

def fmt_inr(v):
    if v is None: return '—'
    try:
        v = float(v)
        if abs(v) >= 1e6:
            return f"₹{v/1e5:.1f}L"
        elif abs(v) >= 1e3:
            return f"₹{v/1e3:.1f}K"
        return f"₹{v:.0f}"
    except: return '—'

def fmt_num(v, decimals=0):
    if v is None: return '—'
    try:
        v = float(v)
        if decimals == 0:
            return f"{v:,.0f}"
        return f"{v:,.{decimals}f}"
    except: return '—'

def fmt_pct(v, already_pct=False, decimals=1):
    if v is None: return '—'
    try:
        v = float(v)
        if not already_pct:
            return f"{v*100:.{decimals}f}%"
        return f"{v:.{decimals}f}%"
    except: return '—'

def parse_scorecard(filepath):
    """Parse the detailed scorecard format.
    Also captures Wk-X'25 columns as last-year (LY) data.

    Returns: (wk_cols, metrics, ly_metrics, ordered_rows)
      metrics:     {metric_name: {wk: value}}  — name-keyed (first occurrence wins)
      ly_metrics:  {metric_name: {wk: ly_value}}
      ordered_rows: list of dicts — one per data row, preserving Excel order
                    {idx, sr_no, name, wk_data: {wk: val}, ly_data: {wk: val}}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        if row[0] == 'Sr. No' and row[1] == 'Metrics':
            header_row = row
            data_start = i + 1
            break

    if not header_row:
        return None, {}, {}, []

    # Identify week columns (Wk-N current) and LY columns (Wk-N'25)
    wk_cols = {}    # wk_num -> col_idx  (current year)
    ly_cols  = {}   # wk_num -> col_idx  (last year, Wk-N'25)
    for j, h in enumerate(header_row):
        if not h:
            continue
        h_str = str(h)
        ly_match = re.match(r"Wk-(\d+)'(\d+)", h_str)
        cur_match = re.match(r'Wk-(\d+)$', h_str)
        if ly_match:
            ly_cols[int(ly_match.group(1))] = j
        elif cur_match:
            wk_cols[int(cur_match.group(1))] = j

    metrics    = {}   # name-keyed (first-seen wins, for KPI/trend sections)
    ly_metrics = {}
    ordered_rows = []

    for idx, row in enumerate(rows[data_start:]):
        metric_name = row[1]
        if not metric_name:
            continue
        wk_data = {wk: safe_float(row[col]) for wk, col in wk_cols.items()}
        ly_data  = {wk: safe_float(row[col]) for wk, col in ly_cols.items()} if ly_cols else {}

        # name-keyed dicts (first occurrence wins to avoid cross-context collisions)
        if metric_name not in metrics:
            metrics[metric_name] = wk_data
        if ly_cols and metric_name not in ly_metrics:
            ly_metrics[metric_name] = ly_data

        ordered_rows.append({
            'idx':     idx,
            'sr_no':   row[0],
            'name':    metric_name,
            'wk_data': wk_data,
            'ly_data': ly_data,
        })

    return wk_cols, metrics, ly_metrics, ordered_rows

def parse_wbr(filepath):
    """Parse the WBR summary format (WK2/WK15)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (Theme, Metric, date1, date2...)
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        if row[0] == 'Theme' and row[1] == 'Metric':
            header_row = row
            data_start = i + 1
            break

    if not header_row:
        return None

    # Date columns → map to week numbers
    date_cols = {}  # col_idx -> week_num
    for j, h in enumerate(header_row):
        if h and re.match(r'\d{4}-\d{2}-\d{2}', str(h)):
            date_str = str(h)[:10]
            # Match to known dates
            for wk, d in WEEK_DATES.items():
                if d == date_str:
                    date_cols[j] = wk
                    break

    metrics = {}
    for row in rows[data_start:]:
        metric_name = row[1]
        if not metric_name:
            continue
        metrics[metric_name] = {}
        for col, wk in date_cols.items():
            metrics[metric_name][wk] = safe_float(row[col])

    return metrics

# ──────────────────────────────────────────────
# BUILD UNIFIED WEEKLY DATA
# ──────────────────────────────────────────────

def build_weekly_data():
    folder = os.path.dirname(os.path.abspath(__file__))
    files = glob(os.path.join(folder, '*.xlsx'))

    # Unified store: metric_name -> {week -> value}  (name-keyed, for KPI/trend)
    unified    = {}
    ly_unified = {}

    # Positional store: row_idx -> {wk -> value}  (position-keyed, for full table)
    # master_rows: authoritative ordered list from the latest file
    positional  = {}   # row_idx -> {wk: value}
    ly_positional = {} # row_idx -> {wk: ly_value}
    master_rows = []   # [{idx, sr_no, name}] from latest (highest-wk) file

    def merge_name(store, metric, wk, val):
        if val is None:
            return
        if metric not in store:
            store[metric] = {}
        if wk not in store[metric]:
            store[metric][wk] = val

    def merge_pos(store, row_idx, wk, val):
        if val is None:
            return
        if row_idx not in store:
            store[row_idx] = {}
        if wk not in store[row_idx]:
            store[row_idx][wk] = val

    # Sort files so latest (highest max-week number) is processed last → becomes master
    file_list = []
    for f in sorted(files):
        if not os.path.basename(f).startswith('~$'):
            file_list.append(f)

    latest_file = None
    latest_max_wk = -1

    for f in file_list:
        wb  = openpyxl.load_workbook(f, data_only=True)
        ws  = wb.active
        hdr = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        max_wk = 0
        for h in hdr:
            if h:
                m = re.match(r'Wk-(\d+)$', str(h))
                if m:
                    max_wk = max(max_wk, int(m.group(1)))
        if max_wk > latest_max_wk:
            latest_max_wk = max_wk
            latest_file   = f

    for f in file_list:
        basename = os.path.basename(f)
        print(f"Reading: {basename}")

        wb  = openpyxl.load_workbook(f, data_only=True)
        ws  = wb.active
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        if first_row[0] and 'WBR Summary' in str(first_row[0]):
            metrics = parse_wbr(f)
            if metrics:
                for m, wk_vals in metrics.items():
                    for wk, val in wk_vals.items():
                        merge_name(unified, m, wk, val)
        else:
            _, metrics, ly_metrics, ordered_rows = parse_scorecard(f)
            for m, wk_vals in metrics.items():
                for wk, val in wk_vals.items():
                    merge_name(unified, m, wk, val)
            for m, wk_vals in ly_metrics.items():
                for wk, val in wk_vals.items():
                    merge_name(ly_unified, m, wk, val)

            # Positional merge
            for row in ordered_rows:
                ri = row['idx']
                for wk, val in row['wk_data'].items():
                    merge_pos(positional, ri, wk, val)
                for wk, val in row['ly_data'].items():
                    merge_pos(ly_positional, ri, wk, val)

            # Set master_rows from the latest file
            if f == latest_file:
                master_rows = ordered_rows

    return unified, ly_unified, positional, ly_positional, master_rows

# ──────────────────────────────────────────────
# TREND ANALYSIS ENGINE
# ──────────────────────────────────────────────

def get_recent_values(unified, metric, last_n=4):
    """Get last N weeks of data for a metric."""
    all_wks = sorted(unified.get(metric, {}).keys())
    vals = [(wk, unified[metric][wk]) for wk in all_wks if unified[metric].get(wk) is not None]
    return vals[-last_n:] if len(vals) >= 1 else vals

def trend_direction(vals):
    """Returns 'up', 'down', 'flat', or 'volatile'."""
    if len(vals) < 2:
        return 'flat'
    changes = [(vals[i][1] - vals[i-1][1]) / abs(vals[i-1][1]) if vals[i-1][1] != 0 else 0 for i in range(1, len(vals))]
    avg_change = sum(changes) / len(changes)
    if all(c > 0.02 for c in changes): return 'up'
    if all(c < -0.02 for c in changes): return 'down'
    if avg_change > 0.05: return 'up'
    if avg_change < -0.05: return 'down'
    return 'flat'

def wow_change(unified, metric):
    vals = get_recent_values(unified, metric, 2)
    if len(vals) < 2:
        return None
    v0, v1 = vals[-2][1], vals[-1][1]
    if v0 == 0:
        return None
    return (v1 - v0) / abs(v0)

# ──────────────────────────────────────────────
# HTML GENERATOR
# ──────────────────────────────────────────────

# ── Format hints keyed by exact metric name as it appears in Excel ──
ROW_FMT = {
    'GMS': 'inr', 'New Buyable GMS': 'inr', 'New Buyable FBA GMS': 'inr',
    'OPS': 'inr', 'OPS to GMS conversion': 'pct', 'ASP': 'inr',
    'Buyable offers': 'num', 'New Buyable offers': 'num', 'New Buyable FBA offers': 'num',
    'FC offers': 'num', 'FTAC offers': 'num', 'Flex offers': 'num',
    'FBA offers': 'num', 'AWAS': 'num', 'FC FTAC': 'num', 'Flex FTAC': 'num',
    'FBA BB GV': 'num', 'FC BB GV': 'num', 'BB GV': 'num',
    'FBA BB GV%': 'pct', 'FC BB GV%': 'pct',
    'Conversion %': 'pct', 'Comp GV': 'num', 'Total GV': 'num',
    'ICPC%': 'pct',
    'RIS Units': 'num', 'Total Units': 'num', 'RIS%': 'pct',
    'IXD IB': 'num', 'FC IB': 'num', 'IXD IB%': 'pct',
    'BxGy Eligible Units': 'num', 'BxGy Promo Units': 'num',
    'BxGy Units coverage': 'pct', 'BxGy Units share': 'pct',
    'Deal OPS': 'inr', 'Coupon OPS': 'inr',
    'Deal OPS%': 'pct', 'Coupon OPS%': 'pct',
    'FC units': 'num', 'FBA units': 'num', 'Total units': 'num',
    'FC units%': 'pct', 'FBA units%': 'pct',
    'Buyable offers T7D': 'num', 'FC offers T7D': 'num', 'FTAC offers T7D': 'num',
    'Flex offers T7D': 'num', 'FBA offers T7D': 'num', 'AWAS T7D': 'num',
    'SCR': 'num',
    'QD GV': 'num', 'QD GV%': 'pct',
    'NCEMI GV': 'num', 'NCEMI%': 'pct',
    'Clicks': 'num',
    'Ad spend': 'inr', 'Attributed sales': 'inr',
    'ACOS': 'pct', 'SP Spend%': 'pct',
    'HT OPS': 'inr', 'MFN HIT OPS': 'inr',
    'FC OOS GV': 'num', 'EF OOS GV': 'num', 'Overall OOS GV': 'num',
    'Overall Instock GV': 'num',
    'FC OOS GV%': 'pct', 'EF OOS GV%': 'pct', 'Overall OOS GV%': 'pct',
}

BAD_HIGH_METRICS = {
    'ACOS', 'FC OOS GV%', 'EF OOS GV%', 'Overall OOS GV%',
    'FC OOS GV', 'EF OOS GV', 'Overall OOS GV', 'SCR',
}

def fmt_cell(val, fmt_type):
    if val is None:
        return '—'
    try:
        f = float(val)
        if fmt_type == 'inr':
            return fmt_inr(f)
        elif fmt_type == 'pct':
            return fmt_pct(f)
        elif fmt_type == 'pct_str':
            return f"{f:.1f}%" if abs(f) < 200 else f"{f:.0f} bps"
        elif fmt_type == 'num':
            return fmt_num(f)
        return str(val)
    except:
        return str(val) if val else '—'

def trend_icon(unified, metric_key, higher_is_good=True):
    vals = get_recent_values(unified, metric_key, 4)
    td = trend_direction(vals)
    if td == 'up':
        color = '#22c55e' if higher_is_good else '#ef4444'
        icon = '▲'
    elif td == 'down':
        color = '#ef4444' if higher_is_good else '#22c55e'
        icon = '▼'
    else:
        color = '#94a3b8'
        icon = '—'
    return f'<span style="color:{color};font-weight:700">{icon}</span>'

def generate_html(unified, ly_unified, positional, ly_positional, master_rows):
    all_weeks = sorted(set(wk for m_vals in unified.values() for wk in m_vals.keys()))
    latest_wk = max(all_weeks) if all_weeks else 15
    latest_date = WEEK_DATES.get(latest_wk, f"Week {latest_wk}")

    # ── KPI cards for latest week ──
    def kpi(label, metric, fmt, higher_is_good=True, prefix=''):
        vals = get_recent_values(unified, metric, 2)
        cur = vals[-1][1] if vals else None
        prev = vals[-2][1] if len(vals) >= 2 else None
        display = fmt_cell(cur, fmt)
        if cur is not None and prev is not None and prev != 0:
            chg = (cur - prev) / abs(prev) * 100
            sign = '+' if chg >= 0 else ''
            chg_color = '#22c55e' if (chg >= 0) == higher_is_good else '#ef4444'
            chg_html = f'<div class="kpi-change" style="color:{chg_color}">{sign}{chg:.1f}% WoW</div>'
        else:
            chg_html = '<div class="kpi-change">—</div>'
        return f'''
        <div class="kpi-card">
          <div class="kpi-label">{label}</div>
          <div class="kpi-value">{display}</div>
          {chg_html}
        </div>'''

    kpi_html = (
        kpi("GMS", "GMS", "inr") or
        kpi("Served GMS", "Served GMS", "inr")
    )
    kpi_html = ""
    for label, metric, fmt, hig in [
        ("GMS", "GMS", "inr", True),
        ("OPS", "OPS", "inr", True),
        ("Served Units", "Served Units", "num", True),
        ("ACOS", "ACOS", "pct", False),
        ("IC Box PC", "ICPC%", "pct", True),
        ("Ad Spend", "Ad spend", "inr", False),
        ("Conversion %", "Conversion %", "pct", True),
        ("Overall OOS %", "Overall OOS GV%", "pct", False),
    ]:
        # try alternate keys
        v = unified.get(metric) or unified.get("Served GMS") if metric == "GMS" else unified.get(metric)
        kpi_html += kpi(label, metric, fmt, hig)

    # ── Build scrollable data table (all rows from Excel, in order) ──
    week_headers = "".join(
        f'<th class="wk-header">Wk {wk}<br><span class="wk-date">{WEEK_DATES.get(wk,"")}</span></th>'
        for wk in all_weeks
    )

    def infer_fmt(name):
        """Infer format type from metric name."""
        return ROW_FMT.get(name, 'num')

    def row_trend_icon(row_idx, name, pos_data):
        """Trend icon based on positional data across weeks."""
        wk_vals = [(wk, pos_data.get(row_idx, {}).get(wk)) for wk in all_weeks]
        wk_vals = [(wk, v) for wk, v in wk_vals if v is not None]
        td = trend_direction(wk_vals[-4:])
        hig = name not in BAD_HIGH_METRICS
        if td == 'up':
            color = '#22c55e' if hig else '#ef4444'; icon = '▲'
        elif td == 'down':
            color = '#ef4444' if hig else '#22c55e'; icon = '▼'
        else:
            color = '#94a3b8'; icon = '—'
        return f'<span style="color:{color};font-weight:700">{icon}</span>'

    table_rows = ""
    for row_meta in master_rows:
        ri      = row_meta['idx']
        sr_no   = row_meta['sr_no']
        name    = row_meta['name']
        fmt     = infer_fmt(name)
        is_sub  = (sr_no is None)   # sub-rows have no Sr. No
        bad_high = name in BAD_HIGH_METRICS

        vals_in_range = [positional.get(ri, {}).get(wk) for wk in all_weeks]

        numeric_vals = [v for v in vals_in_range if v is not None]
        mn = min(numeric_vals) if numeric_vals else 0
        mx = max(numeric_vals) if numeric_vals else 1

        cells = ""
        for wk, val in zip(all_weeks, vals_in_range):
            is_latest = (wk == latest_wk)
            cell_style = "font-weight:700;" if is_latest else ""
            heat = ""
            if val is not None and mx != mn:
                ratio = (val - mn) / (mx - mn)
                if bad_high:
                    ratio = 1 - ratio
                heat = f"background:rgba({255-int(ratio*60)},{175+int(ratio*60)},175,0.25);"
            formatted = fmt_cell(val, fmt)
            cells += f'<td style="{cell_style}{heat}">{formatted}</td>'

        icon = row_trend_icon(ri, name, positional)

        # Visual distinction: Sr. No rows bold, sub-rows indented+lighter
        if is_sub:
            name_style = "padding-left:28px;color:#64748b;font-size:12px;"
        else:
            name_style = "font-weight:600;color:#1e293b;"

        sr_disp = str(int(sr_no)) if sr_no is not None else ""
        table_rows += f'''<tr class="{'sub-row' if is_sub else 'main-row'}">
          <td class="metric-name sticky-col" style="{name_style}">
            {'<span class="sr-badge">'+sr_disp+'</span> ' if sr_disp else ''}{name}
          </td>
          <td class="trend-col">{icon}</td>
          {cells}
        </tr>'''

    # ── Trend narrative ──
    def get_val(metric, wk):
        return unified.get(metric, {}).get(wk)

    def narrative_block(title, color, icon, items):
        items_html = "".join(f"<li>{i}</li>" for i in items if i)
        return f'''<div class="trend-card" style="border-left-color:{color}">
          <div class="trend-title">{icon} {title}</div>
          <ul>{items_html}</ul>
        </div>'''

    # Sales trend
    gms_vals = get_recent_values(unified, 'GMS', 6)
    gms_td = trend_direction(gms_vals)
    ops_vals = get_recent_values(unified, 'OPS', 6)
    asp_vals = get_recent_values(unified, 'ASP', 6)
    units_vals = get_recent_values(unified, 'Served Units', 6)
    acos_vals = get_recent_values(unified, 'ACOS', 6)

    latest_gms = gms_vals[-1][1] if gms_vals else None
    prev_gms = gms_vals[-2][1] if len(gms_vals) >= 2 else None

    sales_items = []
    if latest_gms and prev_gms:
        chg = (latest_gms - prev_gms)/abs(prev_gms)*100
        sales_items.append(f"GMS {fmt_inr(latest_gms)} — {'▲' if chg>0 else '▼'} {abs(chg):.1f}% WoW (prev {fmt_inr(prev_gms)})")
    if asp_vals:
        td = trend_direction(asp_vals)
        sales_items.append(f"ASP trend: {td.upper()} — latest {fmt_inr(asp_vals[-1][1])}")
    if units_vals:
        sales_items.append(f"Served Units latest Wk: {fmt_num(units_vals[-1][1])} | Trend: {trend_direction(units_vals).upper()}")
    if ops_vals and gms_vals:
        conv = ops_vals[-1][1]/gms_vals[-1][1] if gms_vals[-1][1] else None
        if conv:
            sales_items.append(f"OPS/GMS ratio: {conv*100:.1f}% — healthy fulfilment conversion")

    # Operations trend
    oos_vals = get_recent_values(unified, 'Overall OOS GV%', 6)
    fc_oos = get_recent_values(unified, 'FC OOS GV%', 6)
    fba_pct = get_recent_values(unified, 'FBA units%', 6)
    scr_vals = get_recent_values(unified, 'SCR', 6)

    ops_items = []
    if oos_vals:
        td = trend_direction(oos_vals)
        ops_items.append(f"Overall OOS: {fmt_pct(oos_vals[-1][1])} — trend {td.upper()} {'⚠ needs attention' if oos_vals[-1][1] and oos_vals[-1][1]>0.1 else '✓ in control'}")
    if fc_oos:
        ops_items.append(f"FC OOS: {fmt_pct(fc_oos[-1][1])} — {'▼ improving' if trend_direction(fc_oos)=='down' else '▲ worsening' if trend_direction(fc_oos)=='up' else 'stable'}")
    if fba_pct:
        ops_items.append(f"FBA Fulfilment mix: {fmt_pct(fba_pct[-1][1])} — trend {trend_direction(fba_pct).upper()}")
    if scr_vals:
        ops_items.append(f"Seller Cancel Rate (SCR): {fmt_num(scr_vals[-1][1])} (target <5)")

    ris_vals = get_recent_values(unified, 'RIS%', 6)
    if ris_vals:
        ops_items.append(f"RIS%: {fmt_pct(ris_vals[-1][1])} — {'▲ good stock at FC' if ris_vals[-1][1] and ris_vals[-1][1]>0.5 else '▼ low FC stock'}")

    # Marketing trend
    ad_vals = get_recent_values(unified, 'Ad spend', 6)
    acos_td = trend_direction(acos_vals)
    icpc_vals = get_recent_values(unified, 'ICPC%', 6)
    bb_vals = get_recent_values(unified, 'FBA BB GV%', 6)

    mkt_items = []
    if ad_vals:
        mkt_items.append(f"Ad Spend: {fmt_inr(ad_vals[-1][1])} — trend {trend_direction(ad_vals).upper()}")
    if acos_vals:
        mkt_items.append(f"ACOS: {fmt_pct(acos_vals[-1][1])} — trend {acos_td.upper()} {'⚠ above 40%' if acos_vals[-1][1] and acos_vals[-1][1]>0.4 else '✓ efficient'}")
    if icpc_vals:
        mkt_items.append(f"IC Box PC (Price Competitiveness): {fmt_pct(icpc_vals[-1][1])} — {'▲ competitive' if icpc_vals[-1][1] and icpc_vals[-1][1]>0.7 else '▼ non-competitive'}")
    if bb_vals:
        mkt_items.append(f"FBA Buy Box GV%: {fmt_pct(bb_vals[-1][1])} — {'strong' if bb_vals[-1][1] and bb_vals[-1][1]>0.65 else 'moderate'}")

    conv_vals = get_recent_values(unified, 'Conversion %', 6)
    if conv_vals:
        mkt_items.append(f"Conversion Rate: {fmt_pct(conv_vals[-1][1])} — trend {trend_direction(conv_vals).upper()}")

    # Listing trend
    buyable_vals = get_recent_values(unified, 'Buyable offers', 6)
    fba_off = get_recent_values(unified, 'FBA offers', 6)
    awas_vals = get_recent_values(unified, 'AWAS', 6)
    bxgy_vals = get_recent_values(unified, 'BxGy Units coverage', 6)

    lst_items = []
    if buyable_vals:
        lst_items.append(f"Buyable Offers: {fmt_num(buyable_vals[-1][1])} — trend {trend_direction(buyable_vals).upper()} {'⚠ declining' if trend_direction(buyable_vals)=='down' else ''}")
    if fba_off:
        lst_items.append(f"FBA Offers: {fmt_num(fba_off[-1][1])} — listing health {'good' if fba_off[-1][1] and fba_off[-1][1]>55 else 'needs SKU activation'}")
    if awas_vals:
        lst_items.append(f"AWAS: {fmt_num(awas_vals[-1][1])} — active sellers with strong supply")
    if bxgy_vals:
        lst_items.append(f"BxGy Coverage: {fmt_pct(bxgy_vals[-1][1])} — {'strong promo presence' if bxgy_vals[-1][1] and bxgy_vals[-1][1]>0.7 else 'expand BxGy eligible SKUs'}")

    trend_html = f'''
    <div class="trends-grid">
      {narrative_block("Sales Performance", "#6366f1", "📊", sales_items)}
      {narrative_block("Operations & Fulfilment", "#f59e0b", "⚙️", ops_items)}
      {narrative_block("Marketing & Ads", "#ec4899", "📣", mkt_items)}
      {narrative_block("Listing Health", "#10b981", "📋", lst_items)}
    </div>'''

    # ── SWOT ──
    strengths, weaknesses, opportunities, threats = [], [], [], []

    # Strengths
    if gms_vals and len(gms_vals) >= 2 and gms_vals[-1][1] > gms_vals[-2][1]:
        strengths.append("GMS on upward trajectory — strong demand signal")
    if bb_vals and bb_vals[-1][1] and bb_vals[-1][1] > 0.65:
        strengths.append(f"High FBA Buy Box win rate ({fmt_pct(bb_vals[-1][1])}) drives visibility")
    if icpc_vals and icpc_vals[-1][1] and icpc_vals[-1][1] > 0.75:
        strengths.append(f"Strong price competitiveness (IC Box PC {fmt_pct(icpc_vals[-1][1])})")
    if ris_vals and ris_vals[-1][1] and ris_vals[-1][1] > 0.55:
        strengths.append(f"Good FC instock rate (RIS {fmt_pct(ris_vals[-1][1])})")
    if asp_vals and asp_vals[-1][1] and asp_vals[-1][1] > 600:
        strengths.append(f"Premium ASP of {fmt_inr(asp_vals[-1][1])} indicates strong brand positioning")

    # Weaknesses
    if acos_vals and acos_vals[-1][1] and acos_vals[-1][1] > 0.45:
        weaknesses.append(f"High ACOS ({fmt_pct(acos_vals[-1][1])}) — ad spend efficiency needs optimization")
    if oos_vals and oos_vals[-1][1] and oos_vals[-1][1] > 0.08:
        weaknesses.append(f"Overall OOS elevated at {fmt_pct(oos_vals[-1][1])} — lost sales risk")
    if buyable_vals and trend_direction(buyable_vals) == 'down':
        weaknesses.append("Declining buyable offers — catalog suppression or pricing issues")
    if conv_vals and conv_vals[-1][1] and conv_vals[-1][1] < 0.025:
        weaknesses.append(f"Below-average conversion rate ({fmt_pct(conv_vals[-1][1])}) — listing quality or price")

    # Opportunities
    if bxgy_vals and bxgy_vals[-1][1] and bxgy_vals[-1][1] < 0.8:
        opportunities.append("Expand BxGy eligible SKUs to drive basket size and units")
    opportunities.append("Improve keyword targeting and creative to boost GV and conversion")
    if icpc_vals and icpc_vals[-1][1] and icpc_vals[-1][1] < 0.8:
        opportunities.append("Reprice non-competitive SKUs to capture more Buy Box")
    opportunities.append("Launch new SKUs on FBA to grow buyable selection and share")
    if units_vals:
        yoy_target = units_vals[-1][1] * 2.5 if units_vals[-1][1] else None
        if yoy_target:
            opportunities.append(f"Strong YoY growth trend — plan inventory to sustain {fmt_num(yoy_target)} units/wk target")

    # Threats
    threats.append("Rising ad costs (ACOS trend) could erode margin if not managed")
    if oos_vals and trend_direction(oos_vals) == 'up':
        threats.append("OOS trend worsening — supply chain planning required urgently")
    threats.append("Competitor pricing pressure on non-competitive ASINs")
    if scr_vals and scr_vals[-1][1] and scr_vals[-1][1] > 10:
        threats.append(f"High SCR ({fmt_num(scr_vals[-1][1])}) risks account health metrics")

    def swot_col(title, color, bg, items):
        items_html = "".join(f"<li>{i}</li>" for i in items)
        return f'''<div class="swot-box" style="border-top:4px solid {color};background:{bg}">
          <div class="swot-title" style="color:{color}">{title}</div>
          <ul>{items_html}</ul>
        </div>'''

    swot_html = f'''<div class="swot-grid">
      {swot_col("Strengths", "#22c55e", "#f0fdf4", strengths)}
      {swot_col("Weaknesses", "#ef4444", "#fef2f2", weaknesses)}
      {swot_col("Opportunities", "#3b82f6", "#eff6ff", opportunities)}
      {swot_col("Threats", "#f97316", "#fff7ed", threats)}
    </div>'''

    # ── Action Items ──
    p0, p1, p2 = [], [], []

    # P0 — Critical / This week
    if oos_vals and oos_vals[-1][1] and oos_vals[-1][1] > 0.1:
        p0.append(f"Resolve OOS crisis — Overall OOS at {fmt_pct(oos_vals[-1][1])}. Expedite replenishment PO immediately.")
    if acos_vals and acos_vals[-1][1] and acos_vals[-1][1] > 0.5:
        p0.append(f"ACOS at {fmt_pct(acos_vals[-1][1])} — pause non-performing campaigns, review bid strategy.")
    if scr_vals and scr_vals[-1][1] and scr_vals[-1][1] > 15:
        p0.append("SCR critically high — resolve order cancellation root cause to protect account health.")
    if not p0:
        if gms_vals and gms_vals[-1][1]:
            p0.append(f"Monitor GMS trajectory — current at {fmt_inr(gms_vals[-1][1])}. Ensure ad budget is sufficient to maintain momentum.")

    # P1 — High priority / This week
    if acos_vals and acos_vals[-1][1] and 0.35 < acos_vals[-1][1] <= 0.5:
        p1.append(f"Optimize ad spend — ACOS at {fmt_pct(acos_vals[-1][1])}. Review keyword bids and negative keywords.")
    if icpc_vals and icpc_vals[-1][1] and icpc_vals[-1][1] < 0.7:
        p1.append(f"Price audit — IC Box PC at {fmt_pct(icpc_vals[-1][1])}. Identify and reprice non-competitive SKUs.")
    if buyable_vals and trend_direction(buyable_vals) == 'down':
        p1.append(f"Investigate listing suppressions — Buyable offers declining to {fmt_num(buyable_vals[-1][1])}. Resolve any listing quality issues.")
    if not p1:
        p1.append("Review and update A+ content and listing images for top-10 GMS SKUs.")
        p1.append("Audit deals calendar — ensure BxGy / coupon coverage on high-traffic days.")

    # P2 — Normal priority / Next 2 weeks
    p2.append("Plan Q2 inventory buys based on YoY growth trend and current RIS levels.")
    p2.append("Expand FBA selection — target adding 5 new FBA-eligible SKUs to improve buyable selection.")
    p2.append("Set up brand store promotions aligned with upcoming seasonal peaks.")
    if conv_vals and conv_vals[-1][1] and conv_vals[-1][1] < 0.03:
        p2.append(f"Conversion rate ({fmt_pct(conv_vals[-1][1])}) below benchmark — run listing audit, improve main image and bullet points.")

    def action_card(priority, color, bg_color, items):
        items_html = "".join(f'''<li><span class="priority-badge" style="background:{color}">{priority}</span> {i}</li>''' for i in items)
        return f'''<div class="action-card" style="border-left:4px solid {color};background:{bg_color}">
          <div class="action-title" style="color:{color}">{priority} — {'Critical (Resolve Now)' if priority=='P0' else 'High Priority (This Week)' if priority=='P1' else 'Normal Priority (Next 2 Weeks)'}</div>
          <ul class="action-list">{items_html}</ul>
        </div>'''

    actions_html = f'''
    {action_card("P0", "#dc2626", "#fef2f2", p0)}
    {action_card("P1", "#d97706", "#fffbeb", p1)}
    {action_card("P2", "#2563eb", "#eff6ff", p2)}
    '''

    # ── Latest week summary ──
    summary_metrics = [
        ("GMS", "GMS", "inr"),
        ("OPS", "OPS", "inr"),
        ("Served Units", "Served Units", "num"),
        ("Conversion %", "Conversion", "pct"),
        ("ICPC%", "IC Box PC", "pct"),
        ("Ad spend", "Ad Spend", "inr"),
        ("ACOS", "ACOS", "pct"),
        ("Overall OOS GV%", "OOS", "pct"),
        ("Buyable offers", "Buyable Offers", "num"),
        ("FBA BB GV%", "FBA BB GV%", "pct"),
    ]
    summary_cards = ""
    for key, label, fmt in summary_metrics:
        vals = get_recent_values(unified, key, 2)
        cur = vals[-1][1] if vals else None
        prev = vals[-2][1] if len(vals) >= 2 else None
        # LY value for the latest week
        ly_val = ly_unified.get(key, {}).get(latest_wk)
        display = fmt_cell(cur, fmt)

        badges = []
        if cur is not None and prev is not None and prev != 0:
            chg = (cur - prev) / abs(prev) * 100
            sign = '+' if chg >= 0 else ''
            bad_high = key in ('ACOS', 'Overall OOS GV%')
            up_good = (chg >= 0) != bad_high
            chg_color = '#22c55e' if up_good else '#ef4444'
            badges.append(f'<div style="font-size:11px;color:{chg_color};margin-top:2px">{sign}{chg:.1f}% WoW</div>')
        if cur is not None and ly_val is not None and ly_val != 0:
            yoy = (cur - ly_val) / abs(ly_val) * 100
            sign = '+' if yoy >= 0 else ''
            bad_high = key in ('ACOS', 'Overall OOS GV%')
            up_good = (yoy >= 0) != bad_high
            yoy_color = '#22c55e' if up_good else '#ef4444'
            badges.append(f'<div style="font-size:11px;color:{yoy_color};margin-top:2px">{sign}{yoy:.1f}% YoY <span style="color:#94a3b8">(LY: {fmt_cell(ly_val, fmt)})</span></div>')

        summary_cards += f'''<div class="sum-card">
          <div style="font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:0.05em">{label}</div>
          <div style="font-size:18px;font-weight:700;color:#1e293b;margin-top:4px">{display}</div>
          {''.join(badges)}
        </div>'''

    # ── YoY comparison table ──
    yoy_compare_metrics = [
        ("GMS", "GMS", "inr", False),
        ("OPS", "OPS", "inr", False),
        ("Served Units", "Served Units", "num", False),
        ("ASP", "ASP", "inr", False),
        ("Conversion %", "Conversion %", "pct", False),
        ("ICPC%", "IC Box PC %", "pct", False),
        ("Total GV", "Total GV", "num", False),
        ("FBA BB GV%", "FBA BB GV%", "pct", False),
        ("Ad spend", "Ad Spend", "inr", True),
        ("ACOS", "ACOS", "pct", True),
        ("Overall OOS GV%", "Overall OOS %", "pct", True),
        ("Buyable offers", "Buyable Offers", "num", False),
        ("FBA offers", "FBA Offers", "num", False),
        ("Buyable Selection", "Buyable Selection", "num", False),
        ("Prime OPS", "Prime OPS", "inr", False),
        ("Total Deal OPS", "Total Deal OPS", "inr", False),
        ("BxGy OPS", "BxGy OPS", "inr", False),
    ]
    yoy_rows_html = ""
    for key, label, fmt, bad_high in yoy_compare_metrics:
        cur_vals = get_recent_values(unified, key, 1)
        cur = cur_vals[-1][1] if cur_vals else None
        ly  = ly_unified.get(key, {}).get(latest_wk)
        if cur is None and ly is None:
            continue
        cur_disp = fmt_cell(cur, fmt)
        ly_disp  = fmt_cell(ly, fmt)
        if cur is not None and ly is not None and ly != 0:
            yoy = (cur - ly) / abs(ly) * 100
            sign = '+' if yoy >= 0 else ''
            up_good = (yoy >= 0) != bad_high
            yoy_color = '#22c55e' if up_good else '#ef4444'
            arrow = '▲' if yoy >= 0 else '▼'
            yoy_disp = f'<span style="color:{yoy_color};font-weight:700">{arrow} {sign}{yoy:.1f}%</span>'
        else:
            yoy_disp = '—'
        yoy_rows_html += f'''<tr>
          <td style="text-align:left;padding:9px 14px;font-weight:500;color:#374151;border-bottom:1px solid #f1f5f9">{label}</td>
          <td style="text-align:center;padding:9px 14px;font-weight:700;color:#1e293b;border-bottom:1px solid #f1f5f9">{cur_disp}</td>
          <td style="text-align:center;padding:9px 14px;color:#64748b;border-bottom:1px solid #f1f5f9">{ly_disp}</td>
          <td style="text-align:center;padding:9px 14px;border-bottom:1px solid #f1f5f9">{yoy_disp}</td>
        </tr>'''

    yoy_table_html = f'''<div style="background:white;border-radius:10px;padding:20px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,0.06);border:1px solid #e2e8f0;">
      <div style="font-size:14px;font-weight:700;margin-bottom:14px;color:#1e293b">📅 Week {latest_wk} — Year-on-Year Comparison (WK{latest_wk} CY vs WK{latest_wk} LY)</div>
      <div style="overflow-x:auto;">
      <table style="width:100%;border-collapse:collapse;font-size:13px;">
        <thead>
          <tr style="background:#f1f5f9;">
            <th style="text-align:left;padding:10px 14px;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:0.06em;font-weight:700">Metric</th>
            <th style="text-align:center;padding:10px 14px;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:0.06em;font-weight:700">Wk {latest_wk} CY (2026)</th>
            <th style="text-align:center;padding:10px 14px;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:0.06em;font-weight:700">Wk {latest_wk} LY (2025)</th>
            <th style="text-align:center;padding:10px 14px;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:0.06em;font-weight:700">YoY Change</th>
          </tr>
        </thead>
        <tbody>
          {yoy_rows_html}
        </tbody>
      </table>
      </div>
    </div>'''

    # ── Full HTML ──
    now = datetime.now().strftime("%d %b %Y, %H:%M")
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PG Beauty — Amazon MIS | Week {latest_wk}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; background: #f1f5f9; color: #1e293b; }}

  /* HEADER */
  .header {{ background: linear-gradient(135deg, #1e293b 0%, #334155 100%); color: white; padding: 24px 32px; display:flex; justify-content:space-between; align-items:center; }}
  .header-brand {{ font-size: 22px; font-weight: 700; letter-spacing: -0.5px; }}
  .header-sub {{ font-size: 13px; color: #94a3b8; margin-top: 4px; }}
  .header-wk {{ text-align:right; }}
  .wk-badge {{ background: #6366f1; color: white; padding: 6px 14px; border-radius: 20px; font-size: 13px; font-weight: 600; display:inline-block; }}
  .updated {{ font-size: 11px; color: #94a3b8; margin-top: 6px; }}

  /* SECTIONS */
  .section {{ padding: 28px 32px; }}
  .section-title {{ font-size: 18px; font-weight: 700; color: #1e293b; margin-bottom: 16px; padding-bottom: 10px; border-bottom: 2px solid #e2e8f0; display:flex; align-items:center; gap:8px; }}

  /* KPI SUMMARY */
  .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(140px, 1fr)); gap: 12px; margin-bottom: 0; }}
  .sum-card {{ background: white; border-radius: 10px; padding: 14px 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); border: 1px solid #e2e8f0; }}

  /* TABLE */
  .table-wrapper {{ overflow-x: auto; overflow-y: auto; max-height: 600px; border-radius: 12px; box-shadow: 0 1px 8px rgba(0,0,0,0.08); border: 1px solid #e2e8f0; background: white; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 12.5px; }}
  thead th {{ background: #1e293b; color: white; padding: 10px 12px; text-align: center; white-space: nowrap; position: sticky; top: 0; z-index: 10; }}
  thead th.wk-header {{ min-width: 90px; }}
  .wk-date {{ font-size: 10px; color: #94a3b8; font-weight: 400; }}
  .sticky-col {{ position: sticky; left: 0; background: #f8fafc; z-index: 5; font-weight: 500; color: #374151; min-width: 180px; }}
  .trend-col {{ min-width: 40px; text-align:center; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #f1f5f9; text-align: center; white-space: nowrap; }}
  td.metric-name {{ text-align: left; border-right: 1px solid #e2e8f0; }}
  tr.main-row:hover td {{ background: rgba(99,102,241,0.05); }}
  tr.sub-row td {{ border-bottom: 1px solid #f8fafc; }}
  tr.sub-row:hover td {{ background: rgba(99,102,241,0.02); }}
  .group-header td {{ background: #f1f5f9; color: #475569; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; padding: 6px 12px; }}
  .sr-badge {{ display:inline-block; background:#e2e8f0; color:#64748b; font-size:10px; font-weight:700; border-radius:3px; padding:1px 5px; margin-right:4px; }}

  /* TRENDS */
  .trends-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }}
  .trend-card {{ background: white; border-radius: 10px; padding: 20px; border-left: 4px solid; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }}
  .trend-title {{ font-size: 14px; font-weight: 700; margin-bottom: 12px; color: #1e293b; }}
  .trend-card ul {{ list-style: none; }}
  .trend-card li {{ font-size: 13px; color: #475569; padding: 5px 0; border-bottom: 1px solid #f1f5f9; line-height: 1.5; }}
  .trend-card li:last-child {{ border-bottom: none; }}

  /* SWOT */
  .swot-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
  .swot-box {{ border-radius: 10px; padding: 20px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }}
  .swot-title {{ font-size: 14px; font-weight: 700; margin-bottom: 12px; text-transform: uppercase; letter-spacing: 0.05em; }}
  .swot-box ul {{ list-style: none; }}
  .swot-box li {{ font-size: 13px; padding: 5px 0; border-bottom: 1px solid rgba(0,0,0,0.05); padding-left: 14px; position: relative; line-height: 1.5; }}
  .swot-box li::before {{ content: "•"; position: absolute; left: 0; font-weight: 700; }}
  .swot-box li:last-child {{ border-bottom: none; }}

  /* ACTIONS */
  .action-card {{ border-radius: 10px; padding: 20px; margin-bottom: 14px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }}
  .action-title {{ font-size: 14px; font-weight: 700; margin-bottom: 14px; text-transform: uppercase; letter-spacing: 0.05em; }}
  .action-list {{ list-style: none; }}
  .action-list li {{ font-size: 13px; padding: 8px 0; border-bottom: 1px solid rgba(0,0,0,0.06); display: flex; align-items: flex-start; gap: 10px; line-height: 1.5; color: #374151; }}
  .action-list li:last-child {{ border-bottom: none; }}
  .priority-badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; color: white; font-size: 11px; font-weight: 700; flex-shrink: 0; margin-top: 2px; }}

  /* FOOTER */
  .footer {{ text-align: center; padding: 24px; color: #94a3b8; font-size: 12px; border-top: 1px solid #e2e8f0; }}

  @media (max-width: 768px) {{
    .section {{ padding: 16px; }}
    .swot-grid {{ grid-template-columns: 1fr; }}
    .header {{ flex-direction: column; gap: 12px; }}
  }}
</style>
</head>
<body>

<div class="header">
  <div>
    <div class="header-brand">PG Beauty — Amazon MIS Dashboard</div>
    <div class="header-sub">Brand: Parul Garg Beauty &nbsp;|&nbsp; Marketplace: Amazon India &nbsp;|&nbsp; Weekly Business Review</div>
  </div>
  <div class="header-wk">
    <div class="wk-badge">Week {latest_wk} &nbsp;·&nbsp; {latest_date}</div>
    <div class="updated">Updated: {now}</div>
  </div>
</div>

<!-- LATEST WEEK SNAPSHOT -->
<div class="section" style="background:#fff;border-bottom:1px solid #e2e8f0;">
  <div class="section-title">⚡ Week {latest_wk} Snapshot</div>
  <div class="summary-grid">{summary_cards}</div>
</div>

<!-- DATA TABLE -->
<div class="section">
  <div class="section-title">📊 All Metrics — All Weeks (Scroll → and ↓)</div>
  <div class="table-wrapper">
    <table>
      <thead>
        <tr>
          <th class="sticky-col" style="text-align:left;position:sticky;left:0;z-index:15;background:#1e293b">Metric</th>
          <th style="min-width:40px">Trend</th>
          {week_headers}
        </tr>
      </thead>
      <tbody>
        {table_rows}
      </tbody>
    </table>
  </div>
  <div style="font-size:11px;color:#94a3b8;margin-top:8px">
    🟢 Green = relatively better performance &nbsp;|&nbsp; 🔴 Red = relatively weaker &nbsp;|&nbsp; Trend: ▲ rising · ▼ declining · — flat &nbsp;|&nbsp; Heat map is relative within each metric row.
  </div>
</div>

<!-- TREND ANALYSIS -->
<div class="section" style="background:#f8fafc;">
  <div class="section-title">📈 Key Trend Analysis (Last 4–6 Weeks)</div>
  {trend_html}
</div>

<!-- SWOT -->
<div class="section">
  <div class="section-title">🔍 SWOT Analysis</div>
  {swot_html}
</div>

<!-- ACTIONS -->
<div class="section" style="background:#f8fafc;">
  <div class="section-title">✅ Week {latest_wk} — Summary & Action Points</div>
  <div style="background:white;border-radius:10px;padding:20px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,0.06);border:1px solid #e2e8f0;">
    <div style="font-size:14px;font-weight:600;margin-bottom:10px;color:#1e293b">Executive Summary — Week {latest_wk} ({latest_date})</div>
    <p style="font-size:13px;color:#475569;line-height:1.7">
      PG Beauty recorded GMS of <strong>{fmt_inr(gms_vals[-1][1] if gms_vals else None)}</strong> in Week {latest_wk}
      {'with a WoW change of <strong>' + ('+' if (gms_vals[-1][1]-gms_vals[-2][1])>=0 else '') + f"{(gms_vals[-1][1]-gms_vals[-2][1])/abs(gms_vals[-2][1])*100:.1f}%</strong>" if len(gms_vals)>=2 and gms_vals[-2][1] else ''}.
      Served units were <strong>{fmt_num(units_vals[-1][1] if units_vals else None)}</strong>.
      Ad spend stood at <strong>{fmt_inr(ad_vals[-1][1] if ad_vals else None)}</strong> with ACOS of <strong>{fmt_pct(acos_vals[-1][1] if acos_vals else None)}</strong>.
      Buy Box win rate (FBA) is at <strong>{fmt_pct(bb_vals[-1][1] if bb_vals else None)}</strong> and
      overall OOS at <strong>{fmt_pct(oos_vals[-1][1] if oos_vals else None)}</strong>.
      {'Key risk this week is <strong>elevated ACOS</strong> and <strong>OOS</strong> requiring immediate action.' if (acos_vals and acos_vals[-1][1] and acos_vals[-1][1]>0.4) or (oos_vals and oos_vals[-1][1] and oos_vals[-1][1]>0.08) else 'Operations appear stable — focus on scaling GMS through improved conversion and selection expansion.'}
    </p>
  </div>

  {yoy_table_html}

  {actions_html}
</div>

<div class="footer">
  PG Beauty Amazon MIS &nbsp;·&nbsp; Auto-generated from weekly WBR & Scorecard data &nbsp;·&nbsp; {now}
  <br>To update: add new Excel file to the folder and run <code>python3 generate_mis.py</code>
</div>

</body>
</html>'''

    return html

# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

if __name__ == '__main__':
    print("Building unified weekly dataset...")
    unified, ly_unified, positional, ly_positional, master_rows = build_weekly_data()
    print(f"Metrics extracted (name-keyed): {len(unified)}")
    print(f"Weeks covered: {sorted(set(wk for m in unified.values() for wk in m))}")
    print(f"LY metrics available: {len(ly_unified)}")
    print(f"Total table rows (Excel order): {len(master_rows)}")

    print("Generating HTML...")
    html = generate_html(unified, ly_unified, positional, ly_positional, master_rows)

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'amazon_mis.html')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"✓ MIS saved to: {out_path}")
